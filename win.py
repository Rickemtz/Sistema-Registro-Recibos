import pandas as pd
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry
import os
import openpyxl
import win32com.client as win32
import win32print

# Inicializar el DataFrame para almacenar los registros
try:
    df = pd.read_excel('Registros.xlsx')
except FileNotFoundError:
    df = pd.DataFrame(columns=['Número de Socio', 'Nombre'])

# Función para actualizar el DataFrame con un nuevo ingreso o actualización
def actualizar_dataframe(numero_socio, nombre, monto, fecha):
    global df

    monto = float(monto)  # Convertir el monto a tipo float
    if 'Número de Socio' in df.columns and numero_socio in df['Número de Socio'].values:
        idx = df[df['Número de Socio'] == numero_socio].index[0]
        if df.at[idx, 'Nombre'] != nombre:
            messagebox.showerror("Error", "El nombre no coincide con el número de socio existente.")
            return False
        if fecha in df.columns and pd.notna(df.at[idx, fecha]):
            messagebox.showerror("Error", "Ya existe un registro para esta fecha.")
            return False
        df.at[idx, fecha] = monto
    else:
        nuevo_registro = pd.DataFrame({'Número de Socio': [numero_socio], 'Nombre': [nombre], fecha: [monto]})
        df = pd.concat([df, nuevo_registro], ignore_index=True)
    return True

# Función para guardar el DataFrame actualizado en un archivo Excel
def guardar_en_excel():
    try:
        df.to_excel('Registros.xlsx', index=False)
        return True
    except PermissionError:
        messagebox.showerror("Error de permisos", "No tienes permisos para guardar en el archivo Excel. Verifica los permisos de escritura.")
        return False
    except Exception as e:
        messagebox.showerror("Error al guardar", f"Ocurrió un error al guardar en el archivo Excel: {e}")
        return False

# Función para imprimir con Excel y enviar comando de corte
def imprimir_con_excel(texto_recibo):
    nombre_archivo = 'Recibo.xlsx'

    # Eliminar archivo si ya existe para evitar conflictos
    if os.path.exists(nombre_archivo):
        try:
            os.remove(nombre_archivo)
        except Exception as e:
            messagebox.showerror("Error de archivo", f"No se pudo eliminar el archivo existente: {e}")
            return

    # Crear y guardar el archivo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Escribir el recibo en las celdas A1:A8
    recibo_texto = texto_recibo.split('\n')
    for i, linea in enumerate(recibo_texto):
        ws.cell(row=i+1, column=1, value=linea)

    # Ajustar el ancho de las columnas para que se ajusten al texto
    ws.column_dimensions['A'].width = 40

    # Ajustar el formato del texto y centrar los datos
    for row in ws.iter_rows(min_row=1, max_row=len(recibo_texto), min_col=1, max_col=1):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Configurar los márgenes
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    # Guardar el archivo antes de imprimir
    wb.save(nombre_archivo)

    try:
        # Iniciar Excel
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False

        # Abrir el archivo de Excel
        wb = excel.Workbooks.Open(os.path.abspath(nombre_archivo))

        # Imprimir el archivo
        wb.PrintOut()
        messagebox.showinfo("Impresión", "¡Recibo generado con éxito!")

        # Cerrar el archivo sin guardar cambios
        wb.Close(SaveChanges=False)
        excel.Quit()

    except Exception as e:
        messagebox.showerror("Error de impresión", f"Error al generar el recibo: {e}")
    finally:
        # Asegurarse de que Excel se cierre
        if 'excel' in locals():
            excel.Quit()

        # Eliminar el archivo de Excel después de imprimir
        if os.path.exists(nombre_archivo):
            os.remove(nombre_archivo)

# Función para imprimir un recibo
def imprimir_registro(numero_socio, nombre, monto, fecha):
    texto_recibo = (
        f"RECIBO\n"
        f"                                     \n"
        f"*************************************\n"
        f"                                     \n"
        f"Recibos por Aportación               \n"
        f"                                     \n"
        f"Socio: {numero_socio} {nombre}       \n"
        f"${monto:.2f}      {fecha}            \n"
        f"*************************************\n"
        f"                                     \n"
        f"GRACIAS POR SU COLABORACIÓN          \n"
        f"                                     \n"
        f"*************************************\n"
        f"                                     \n"
        f"                                     \n"
        f"                                     \n"
        f"                                     \n"
    )

    recibos_text.config(state=tk.NORMAL)
    recibos_text.delete(1.0, tk.END)
    recibos_text.insert(tk.END, texto_recibo)
    recibos_text.config(state=tk.DISABLED)

# Función para registrar un nuevo ingreso
def registrar():
    try:
        numero_socio = int(numero_socio_entry.get())
        nombre = nombre_entry.get().strip()
        monto = float(monto_entry.get())
        fecha_seleccionada_value = fecha_seleccionada.get_date().strftime("%d/%m/%Y")

        if actualizar_dataframe(numero_socio, nombre, monto, fecha_seleccionada_value):
            if guardar_en_excel():
                imprimir_registro(numero_socio, nombre, monto, fecha_seleccionada_value)
                messagebox.showinfo("Registro", "¡Registro exitoso!")
            else:
                # Revertir el DataFrame si no se pudo guardar el archivo
                df.drop(df.tail(1).index, inplace=True)
        else:
            # No generar recibo si no se puede actualizar el DataFrame
            recibos_text.config(state=tk.NORMAL)
            recibos_text.delete(1.0, tk.END)
            recibos_text.config(state=tk.DISABLED)

        numero_socio_entry.delete(0, tk.END)
        nombre_entry.config(state='normal')
        nombre_entry.delete(0, tk.END)
        monto_entry.delete(0, tk.END)
    except ValueError as e:
        messagebox.showerror("Error", f"Por favor, ingrese valores válidos: {e}")

# Función para buscar y cargar información de un socio
def buscar_socio():
    try:
        numero_socio = int(numero_socio_entry.get())
        if 'Número de Socio' in df.columns and numero_socio in df['Número de Socio'].values:
            socio_info = df[df['Número de Socio'] == numero_socio].iloc[0]
            nombre_entry.config(state='normal')
            nombre_entry.delete(0, tk.END)
            nombre_entry.insert(0, socio_info['Nombre'])
            nombre_entry.config(state='readonly')

            # Obtener las fechas disponibles para el socio con valores no nulos
            fechas = [col for col in socio_info.index if col not in ['Número de Socio', 'Nombre'] and pd.notna(socio_info[col])]
            fechas_disponibles.set('')
            fechas_combobox['values'] = fechas
        else:
            messagebox.showinfo("Información", f"No se encontró información para el Número de Socio {numero_socio}.")
    except ValueError:
        messagebox.showerror("Error", "Ingrese un número de socio válido.")

# Función para imprimir el último registro
def imprimir_ultimo_registro():
    texto_recibo = recibos_text.get("1.0", tk.END)
    if texto_recibo.strip():
        imprimir_con_excel(texto_recibo)
    else:
        messagebox.showerror("Error", "No hay recibo para imprimir.")

# Función para mostrar el recibo de la fecha seleccionada
def mostrar_recibo_fecha_seleccionada(event):
    try:
        numero_socio = int(numero_socio_entry.get())
        fecha = fechas_combobox.get()
        if 'Número de Socio' in df.columns and numero_socio in df['Número de Socio'].values:
            socio_info = df[df['Número de Socio'] == numero_socio].iloc[0]
            nombre = socio_info['Nombre']
            monto = socio_info.get(fecha, 0.0)

            texto_recibo = (
                f"RECIBO\n"
                f"                                     \n"
                f"*************************************\n"
                f"                                     \n"
                f"Recibos por Aportación               \n"
                f"                                     \n"
                f"Socio: {numero_socio} {nombre}       \n"
                f"${monto:.2f}      {fecha}            \n"
                f"*************************************\n"
                f"                                     \n"
                f"GRACIAS POR SU COLABORACIÓN          \n"
                f"                                     \n"
                f"*************************************\n"
                f"                                     \n"
                f"                                     \n"
                f"                                     \n"
            )

            recibos_text.config(state=tk.NORMAL)
            recibos_text.delete(1.0, tk.END)
            recibos_text.insert(tk.END, texto_recibo)
            recibos_text.config(state=tk.DISABLED)
    except ValueError:
        messagebox.showerror("Error", "Por favor, ingrese un número de socio válido y seleccione una fecha.")

# Función para limpiar la pantalla
def limpiar_pantalla():
    numero_socio_entry.delete(0, tk.END)
    nombre_entry.config(state='normal')
    nombre_entry.delete(0, tk.END)
    monto_entry.delete(0, tk.END)
    fecha_seleccionada.set_date(None)
    fechas_disponibles.set('')
    fechas_combobox['values'] = []
    recibos_text.config(state=tk.NORMAL)
    recibos_text.delete(1.0, tk.END)
    recibos_text.config(state=tk.DISABLED)

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Sistema de Registro y Generación de Recibos")

tk.Label(root, text="Número de Socio:").grid(row=0, column=0)
numero_socio_entry = tk.Entry(root)
numero_socio_entry.grid(row=0, column=1)

tk.Label(root, text="Nombre:").grid(row=1, column=0)
nombre_entry = tk.Entry(root)
nombre_entry.grid(row=1, column=1)

tk.Label(root, text="Monto de ingreso:").grid(row=2, column=0)
monto_entry = tk.Entry(root)
monto_entry.grid(row=2, column=1)

tk.Label(root, text="Fecha:").grid(row=3, column=0)
fecha_seleccionada = DateEntry(root, width=18, background='darkblue', foreground='white', borderwidth=2)
fecha_seleccionada.grid(row=3, column=1)

tk.Label(root, text="Fechas registradas:").grid(row=4, column=0)
fechas_disponibles = tk.StringVar()
fechas_combobox = ttk.Combobox(root, textvariable=fechas_disponibles)
fechas_combobox.grid(row=4, column=1)
fechas_combobox.bind("<<ComboboxSelected>>", mostrar_recibo_fecha_seleccionada)

registrar_button = tk.Button(root, text="Registrar", command=registrar)
registrar_button.grid(row=5, column=0, columnspan=2, pady=10)

recibos_text = tk.Text(root, height=20, width=50)
recibos_text.grid(row=6, column=0, columnspan=2)
recibos_text.config(state=tk.DISABLED)

buscar_button = tk.Button(root, text="Buscar Socio", command=buscar_socio)
buscar_button.grid(row=0, column=2, pady=10)

imprimir_button = tk.Button(root, text="Imprimir", command=imprimir_ultimo_registro)
imprimir_button.grid(row=5, column=2, pady=10)

limpiar_button = tk.Button(root, text="Limpiar", command=limpiar_pantalla)
limpiar_button.grid(row=5, column=3, pady=10)

root.mainloop()
